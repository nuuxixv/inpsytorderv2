# -*- coding: utf-8 -*-
"""
검사 위계 빌드 툴 (v1)
입력 : 상품_목록.xlsx
출력 : 검사군마스터.xlsx + 상품매핑.xlsx  (= DB 시드 원본)
구조 : ① 자동 레이어(첫단어 묶기 + LCP 검사명 + 옵션명)  ② 특이건 OVERRIDES
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict

SRC = r"C:\Users\김건우\Downloads\상품_목록.xlsx"
OUT_MASTER = r"C:\Users\김건우\Downloads\검사군마스터.xlsx"
OUT_MAP = r"C:\Users\김건우\Downloads\상품매핑.xlsx"

# ───────── 특이건 오버라이드 (16개를 데이터로) ─────────
OVERRIDES = {
    'K-CDI':  {'split': [('한국어판 아동우울척도 2판', '아동우울척도'),
                         ('영아발달검사',             '영아발달'),
                         ('아동발달검사',             '아동발달')]},
    'STROOP': {'split': [('노인용 한국판 스트룹 색상-단어 검사', '노인용'),
                         ('스트룹 아동 색상-단어검사',          '아동')]},
    # split 항목: (검사명, 매칭키워드, [약어]) — 약어 생략 시 첫단어 사용
    '한국판': {'split': [('Young 심리도식 질문지', 'Young'),
                         ('간이정신상태검사 2판', '간이정신상태', 'K-MMSE-2')]},
    # SCT는 검색 편의로 3개(SCT/SCT-A/SCT-C) 유지 — 공용 지침서는 SHARED_OPTIONS로 복제 노출
    'PAI-A':  {'name': '청소년 성격평가 질문지'},
    'CogMCI': {'name': '인지중재치료프로그램'},
    'NAS-PI': {'name': '다면적 분노검사'},
    'K-CSBS': {'abbr': 'K-CSBS DP', 'name': '한국판 의사소통 및 상징행동 발달 검사'},
    'K-CTC':  {'delete': True},
    'K-MBRS': {'delete': True},
}

# 병합 대상(피흡수 첫단어)은 대표가 흡수하므로 단독 처리 금지
MERGE_MEMBERS = set()
for _f, _ov in OVERRIDES.items():
    MERGE_MEMBERS.update(_ov.get('merge', []))

# 공용 옵션 복제 노출: 한 상품을 여러 검사군 2뎁스에 표시 (주문은 동일 product)
SHARED_OPTIONS = [
    # A안: 검사별 '(공용)(1)' 옵션이 이미 각 검사군에 있어 복제 불필요.
    # 공용 옵션 복제가 필요한 검사가 새로 생기면 여기에 추가.
]

# 검사 위계에서 제외할 상품 (중복 등록된 도서 단품 등)
EXCLUDE_CODES = {'PITM000132_1'}  # '이해와 활용' 도서 — 검사별 (공용)(1)과 중복

# 옵션명 개별 예외 (규칙으로 처리 불가한 케이스)
OPTION_OVERRIDE = {
    'PITM000090_1': '자극카드(16장)',   # (16장)=내용 vs (1)=묶음수량 충돌 → (1) 제외
}

def lcp(strs):
    if not strs: return ''
    a = min(strs); b = max(strs); i = 0
    while i < len(a) and i < len(b) and a[i] == b[i]: i += 1
    return a[:i]

# 연령/월령/대상 표현 (옵션명에서 빼서 말머리로 — 옵션명엔 형태만 남김)
AT_PAT = re.compile(r'(\d+\s*~\s*\d+\s*개월용?|\d+\s*개월용?|\d+\s*-\s*\d+\s*세용?|\d+\s*세용?|초등\s*고학년|초등\s*저학년|고학년|저학년|초등\s*\d\s*학년|\d\s*학년|만\s*\d+[\s~\d]*세|대학생용|성인기용|성인용|아동용|유아용|노인용|청소년용|부모용|교사용|영아용\s*\d?|자기보고용?|[가-힣]+\s*보고형|정보제공자)')

# ───────── 읽기 ─────────
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
rows = list(wb.active.iter_rows(values_only=True))
NAME, CODE, CAT = 0, 1, 2
products = []
for r in rows[1:]:
    if not r or r[NAME] is None: continue
    name = re.sub(r'\s+', ' ', str(r[NAME])).strip()   # 연속공백 정규화
    code = str(r[CODE]).strip() if r[CODE] is not None else ''
    cat = str(r[CAT]).strip() if len(r) > CAT and r[CAT] is not None else ''
    if not name: continue
    if code in EXCLUDE_CODES: continue
    products.append((name, code, cat))

# category 분포
catdist = defaultdict(int)
for n, c, ct in products: catdist[ct] += 1
print("카테고리 분포:", dict(catdist))

# ───────── 첫단어 그룹 ─────────
groups = defaultdict(list)
for name, code, cat in products:
    groups[name.split()[0]].append((name, code, cat))

# ───────── 검사군 빌드 ─────────
test_groups = []   # {abbr, name, items}
processed = set()
deleted = []

for first, items in groups.items():
    if first in processed or first in MERGE_MEMBERS: continue
    ov = OVERRIDES.get(first, {})
    if ov.get('delete'):
        deleted.append((first, len(items))); continue
    if 'merge' in ov:
        merged = list(items)
        for m in ov['merge']:
            merged += groups.get(m, []); processed.add(m)
        names = [n for n, c, ct in merged]
        test_groups.append({'abbr': ov.get('abbr', first),
                             'name': ov.get('name') or lcp(names),
                             'items': merged})
    elif 'split' in ov:
        matched = set()
        for spec in ov['split']:
            sub_name, kw = spec[0], spec[1]
            sub_abbr = spec[2] if len(spec) > 2 else first
            sub = [(n, c, ct) for n, c, ct in items if kw in n]
            for n, c, ct in sub: matched.add(c)
            if sub:
                test_groups.append({'abbr': sub_abbr, 'name': sub_name, 'items': sub})
        left = [(n, c, ct) for n, c, ct in items if c not in matched]
        if left:
            test_groups.append({'abbr': first, 'name': f'⚠미매칭({first})', 'items': left})
    else:
        names = [n for n, c, ct in items]
        abbr = ov.get('abbr', first)
        paren_start = abbr.startswith('(')        # 괄호로 시작 = 약어 무의미
        if paren_start: abbr = ''
        if ov.get('name'):
            name = ov['name']
        elif len(items) > 1:
            pre = lcp(names)
            name = pre.strip(' _·-:') if paren_start else (pre[len(first):].strip(' _·-:') or pre)
        else:
            body = (names[0] if paren_start else names[0][len(first):]).strip(' _·-:')
            name = body.split('_')[0].strip() if '_' in body else body   # 첫 "_" 앞 = 검사명
        test_groups.append({'abbr': abbr, 'name': name, 'items': items})

# ───────── 옵션명 · 말머리 · 공용 ─────────
def options_of(grp):
    items = grp['items']
    names = [n for n, c, ct in items]
    pre = lcp(names) if len(items) > 1 else grp['abbr']
    out = []
    for idx, (n, c, ct) in enumerate(sorted(items), 1):
        if len(items) > 1:
            opt = n[len(pre):].strip(' _/')
        else:                                        # 옵션1개: 약어 뗀 뒤 첫 "_" 뒤 = 옵션명
            body = n[len(grp['abbr']):].strip(' _/')
            parts = body.split('_', 1)
            opt = parts[1].strip() if len(parts) > 1 else body
        if not opt: opt = n
        opt = re.sub(r'\(\s*[^():]*:\s*([^()]*?)\s*\)', r'\1', opt).strip()   # "(약어: 형태)" → "형태"
        # 옵션명에 검사명이 중복되면 제거 (말머리가 검사명 앞에 붙는 CogMCI류)
        if grp['name'] and grp['name'] in opt:
            opt = re.sub(r'\s+', ' ', opt.replace(grp['name'], ' ')).strip(' _/')
        # 말머리 분리 → 옵션명엔 형태만 남김 (말머리/옵션명 중복 방지)
        mp = re.match(r'([^()_]+)\)\s*_?\s*', opt)   # "보고형)_..." 괄호조각
        if mp:
            head = mp.group(1).strip()
            opt = opt[mp.end():].strip(' _/')
        elif '_' in opt:                              # "말머리_형태" → 말머리 떼고 형태만
            head, opt = opt.split('_', 1)
            head, opt = head.strip(), opt.strip(' _/')
        else:
            head = ''
        # 옵션명에 남은 연령/월령/대상 → 말머리로 (단 "(...공용...)" 괄호는 지침서 적용범위라 보존)
        guard = re.findall(r'\([^()]*공용[^()]*\)', opt)
        masked = opt
        for gi, gv in enumerate(guard):
            masked = masked.replace(gv, f'\x00{gi}\x00', 1)
        found = AT_PAT.findall(masked)
        if found:
            for f in found:
                masked = masked.replace(f, '')
            masked = re.sub(r'[\s:_]*([(/])', r'\1', masked)
            masked = re.sub(r'\s+', ' ', masked).strip(' _/:-')
            extra = ' '.join(dict.fromkeys(found))
            head = (head + ' ' + extra).strip() if head else extra
        for gi, gv in enumerate(guard):
            masked = masked.replace(f'\x00{gi}\x00', gv)
        opt = masked
        common = 'Y' if '공용' in n else ''
        opt = OPTION_OVERRIDE.get(c, opt)        # 개별 옵션명 예외
        out.append((c, n, opt, head, common, ct, idx))
    for (it, label) in grp.get('shared', []):
        n, c, ct = it
        out.append((c, n, label, '공용지침서', 'Y', ct, len(out) + 1))
    return out

# 공용 옵션 복제 노출 (LCP 오염 방지 위해 본 items에서 빼고 shared로 분리)
for g in test_groups:
    g['shared'] = []
code2item = {c: (n, c, ct) for n, c, ct in products}
for sh in SHARED_OPTIONS:
    it = code2item.get(sh['code'])
    if not it:
        print(f"⚠ SHARED 코드 없음: {sh['code']}"); continue
    for g in test_groups:
        g['items'] = [x for x in g['items'] if x[1] != sh['code']]
        if g['abbr'] in sh['groups']:
            g['shared'].append((it, sh['label']))

# 검사명 끝 괄호 정리 (LCP가 옵션 공통 괄호까지 먹은 경우)
for g in test_groups:
    nm = re.sub(r'\s*\([^()]*\)\s*$', '', g['name']).strip()   # 끝 닫힌괄호
    nm = re.sub(r'\s*\(\s*$', '', nm).strip()                  # 끝 열린괄호 "( "
    if nm:
        g['name'] = nm

# 정렬: 카테고리 → 검사명
test_groups = [g for g in test_groups if g['items'] or g['shared']]
test_groups.sort(key=lambda g: (g['items'][0][2] if g['items'] else g['shared'][0][0][2], g['name']))

# ───────── 출력 1: 검사군마스터 ─────────
mwb = openpyxl.Workbook(); ms = mwb.active; ms.title = '검사군마스터'
ms.append(['정렬', '약어', '검사명', '옵션수', '카테고리'])
for c in ms[1]: c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='DDDDDD')
warn = PatternFill('solid', fgColor='FFE0E0')
for i, g in enumerate(test_groups, 1):
    ms.append([i, g['abbr'], g['name'], len(g['items']), g['items'][0][2]])
    if '⚠' in g['name']:
        for cell in ms[ms.max_row]: cell.fill = warn
for i, w in enumerate([6, 16, 36, 7, 10], 1):
    ms.column_dimensions[get_column_letter(i)].width = w
ms.freeze_panes = 'A2'
mwb.save(OUT_MASTER)

# ───────── 출력 2: 상품매핑 (DB 시드 원본) ─────────
pwb = openpyxl.Workbook(); ps = pwb.active; ps.title = '상품매핑'
ps.append(['검사군약어', '검사명', '옵션명', '말머리', '공용', '옵션정렬', '상품코드', '상품명(원본)', '카테고리'])
for c in ps[1]: c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='DDDDDD')
for g in test_groups:
    for (code, name, opt, head, common, cat, idx) in options_of(g):
        ps.append([g['abbr'], g['name'], opt, head, common, idx, code, name, cat])
for i, w in enumerate([15, 30, 38, 18, 6, 7, 20, 50, 9], 1):
    ps.column_dimensions[get_column_letter(i)].width = w
ps.freeze_panes = 'A2'
pwb.save(OUT_MAP)

# ───────── 요약 ─────────
print(f"검사군: {len(test_groups)}개 / 삭제(절판): {deleted}")
print(f"저장: {OUT_MASTER}")
print(f"      {OUT_MAP}")
unmatched = [g['name'] for g in test_groups if '⚠' in g['name']]
print(f"⚠ 미매칭 검사군: {unmatched if unmatched else '없음'}")
# 말머리 자동 채움률
total_opt = sum(len(g['items']) for g in test_groups)
with_head = sum(1 for g in test_groups for o in options_of(g) if o[3])
print(f"옵션 {total_opt}개 중 말머리 자동 채움: {with_head}개")
print("=== 보정 검사군 검증 샘플 ===")
for g in test_groups:
    if g['abbr'] in ('K-CDI','SCT','K-MMSE-2','PAI-A','CogMCI','NAS-PI','K-CSBS DP') or '문장완성' in g['name'] or 'Young' in g['name'] or '간이정신' in g['name']:
        print(f"  [{g['abbr']}] {g['name']} (옵션 {len(g['items'])})")
print("=== 옵션·말머리 샘플 ===")
for g in test_groups:
    if g['name'] in ('간이정신상태검사 2판', '청소년 성격평가 질문지', '인지중재치료프로그램') or '문장완성' in g['name'] or '걸음마기' in g['name'] or '자아개념' in g['name']:
        print(f"[{g['abbr']}] {g['name']}")
        for o in options_of(g)[:8]:
            print(f"   말머리=「{o[3]}」  공용={o[4] or '-'}  옵션명={o[2]}")
